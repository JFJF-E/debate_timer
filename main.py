import os
import time
import threading
from mimetypes import init
import sys
from PyQt5.QtWidgets import QApplication, QMainWindow
from openpyxl import load_workbook
import timer
# import RPi.GPIO


time1=''
time2=''
state_txt = '阶段1'
class mainui:
    def __init__(self,buttom1_state=1,buttom2_state=0,txt1='',txt2=''):
        global buttom_pal,buttom_stop,time1,time2,state_txt
        self.buttom1_state = buttom1_state
        self.buttom2_state = buttom2_state
        self.txt1 = txt1
        self.txt2 = txt2
        self.app = QApplication(sys.argv)
        MainWindow = QMainWindow()
        self.ui = timer.Ui_MainWindow()
        self.ui.setupUi(MainWindow)
        MainWindow.show()
        
        # ui.label.setText.connect(time1)
        self.ui.label.setText(time1)
        self.ui.label_2.setText(time2)
        self.ui.pushButton.clicked.connect(self.wr_bt1)
        self.ui.pushButton_2.clicked.connect(self.wr_bt2)
        
        t1 = threading.Thread(target=self.fresh_txt)
        t1.start()


        self.buttom2_state=0
        sys.exit(self.app.exec_())
    def fresh_txt(self):
        while(True):
            self.ui.label_3.setText(state_txt)
            self.ui.label.setText(time1)
            self.ui.label_2.setText(time2)
            time.sleep(0.1)
        
            # print('刷新成功')
    def start(self):
        print("ok start")
    def check_buttom1(self):
        if(self.buttom1_state == 1):
            self.buttom1_state=0
            return 1
        
    def check_buttom2(self):
        if(self.buttom2_state == 1):
            self.buttom2_state=0
            return 1
    def wr_bt1(self):
        global buttom_pal
        self.buttom1_state = 1
        print(10)
        buttom_pal=1
    def wr_bt2(self):
        global buttom_stop
        self.buttom2_state = 1
        print(20)
        buttom_stop=1




class TT(threading.Thread):
    def run(self) -> None:
        mainui()

T2 = TT()
T2.start()



print('ok')
local_path = os.getcwd()
wb = load_workbook(local_path+"/时间.xlsx")

def sensor(team,option="defalt"):
    ret = None
    if(option == "defalt"):
        pass
    return ret

def ticktime(option):
    global ticks1,ticks2,time_start
    if(option == "check"):
        pass
    elif(option == "delta"):
        delta =0.1- time.time()+time_start
        print('sleep')
        return delta
    elif(option == "freeze"):
        pass
def display(arg,option):
    if(option == ""):
        pass

def check_buttom():
    global buttom_pal,buttom_stop
    pass
        

def exit_check():
    global state,ticks1,ticks2,maxtick,beepflag1,beepflag2
    if(buttom_stop == 1):
        # state+=1
        return False
    elif(state!=7 and (ticks1>maxtick or ticks2>maxtick)):
        # state+=1
        return False
    
    #自由辩论
    elif(state==7):
        #1
        if(ticks1>maxtick-300):
            if(beepflag1==0):
                warnlist[0]=1
                beepflag1=1
            if(ticks1>maxtick):warnlist[1]=1
        #2
        if(ticks2>maxtick-300):
            if(beepflag2==0):
                warnlist[2]=1
                beepflag2=1
            if(ticks2>maxtick):warnlist[3]=1
    elif(ticks1>maxtick and ticks2>maxtick):
        # state+=1
        return False
    return True

def load_option(state):
    return ticks_list[state-1]

def warn_stop(warnlist):
    pass

def pal_or_break():
    global buttom_pal,buttom_stop,time_start,state_txt
    if(buttom_pal==1):
        buttom_pal=0
        # state_txt =state_list[state]+'(暂停状态)'
        while(buttom_pal==0 and buttom_stop==0):
            # check_buttom()
            time.sleep(0.2)
            # print('等待再次暂停，buttom_pal:',buttom_pal)
        buttom_pal=0
        # state_txt = state_list[state]
        time_start=time.time()
    if(buttom_stop==1):
        buttom_stop=0
        return 1
    
    return 0

def time_formater(sec_in):
    if(sec_in>600):
        return str(sec_in//600)+':'+str((sec_in%600)//10)+'.'+str(sec_in%10)+'s'
    else:
        return '0 :'+str((sec_in%600)//10)+'.'+str(sec_in%10)+'s'


state = 1
state_list = ['','左方一辩','右方一辩','左方二辩','右方二辩','左方三辩','右方三辩','自由辩论','右方四辩','左方四辩']
ticks_list=[]
warnlist=[0,0,0,0]

ticks1=0
ticks2=0

maxtick=0

beepflag1=0
beepflag2=0

buttom_pal=0
buttom_stop=0

for ws in wb:
    for val in range(len(ws['B2':'B6'])):
        cell = ws['B2':'B6'][val]
        ticks_list.append(cell[0].value)
        if(val!=3):
            ticks_list.append(ticks_list[-1])
    print(ticks_list)

time.sleep(0.2)
for i in range(len(ticks_list)):
    buttom_pal = 1
    print('state',state)
    ticks1=0
    ticks2=0
    maxtick= load_option(state)
    warnlist=[0,0,0,0]
    beepflag1=0
    beepflag2=0
    state_txt = state_list[state]
    while(exit_check()):
        time_start = time.time()
        # check_buttom()
        if(pal_or_break()==1):
            break
        t1 = sensor(1)
        t2 = sensor(2)
        if(state in [1,3,5,9]):
            time1 =time_formater(maxtick - ticks1)
            print(time1)
            ticks1+=1
            time.sleep(ticktime("delta"))
        elif(state in [2,4,6,8]):
            time2 =time_formater(maxtick - ticks2)
            ticks2+=1
            time.sleep(ticktime("delta"))
        elif(state ==7):
            if(sensor(1)):
                time1 =time_formater(maxtick - ticks1)
                ticks1+=1
            if(sensor(2)):
                time2 =time_formater(maxtick - ticks2)
                ticks2+=1
            time.sleep(ticktime("delta"))
    state+=1
    
