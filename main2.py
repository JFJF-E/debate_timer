import os
import time
import threading
from mimetypes import init
import sys
from PyQt5.QtWidgets import QApplication, QMainWindow
from openpyxl import load_workbook
import timer
import RPi.GPIO as GPIO

'''
gpio_light_right= 1
gpio_light_left=2
gpio_beeper_right=3
gpio_beeper_left=4
'''
gpio_beeper = 21
gpio_buttom = 20

phy_but = 0



#设置 GPIO 模式为 BCM
GPIO.setmode(GPIO.BCM)  
#定义 GPIO 引脚
GPIO.setup(gpio_beeper,GPIO.OUT)
GPIO.setup(gpio_buttom,GPIO.IN)
#设置 GPIO 的工作方式 (IN / OUT)
'''
GPIO.setup(GPIO_TRIGGER1, GPIO.OUT)
GPIO.setup(GPIO_ECHO1, GPIO.IN)
GPIO.setup(GPIO_TRIGGER2, GPIO.OUT)
GPIO.setup(GPIO_ECHO2, GPIO.IN)
'''


warn0 = 0
warn1 =   0
warn2= 0
warn3   = 0


time1=''
time2=''
state_txt = ''


class mainui:
    """
    简单的ui，极其简单的ui  基于pyqt5
    """
    def __init__(self,buttom1_state=1,buttom2_state=0,txt1='',txt2=''):
        global buttom_pal,buttom_stop,time1,time2,state_txt

        self.buttom1_state = buttom1_state
        self.buttom2_state = buttom2_state
        self.txt1 = txt1
        self.txt2 = txt2
        self.app = QApplication(sys.argv)
        MainWindow = QMainWindow()
        self.ui = timer.Ui_MainWindow()
        self.ui.setupUi(MainWindow)
        MainWindow.show()
        
        # ui.label.setText.connect(time1)
        self.ui.label.setText(time1)
        self.ui.label_2.setText(time2)

        #设置软件按钮
        self.ui.pushButton.clicked.connect(self.wr_bt1)
        self.ui.pushButton_2.clicked.connect(self.wr_bt2)
        
        #ui 字符刷新线程  采取一直读取全局变量的方法刷新
        t1 = threading.Thread(target=self.fresh_txt)
        t1.start()


        self.buttom2_state=0
        sys.exit(self.app.exec_())
    def fresh_txt(self):
        #刷新字体的函数  在单独线程里面运行 直接死循环
        while(True):
            try:
                self.ui.label_3.setText(state_txt)
                self.ui.label.setText(time1)
                self.ui.label_2.setText(time2)
                # print(distance())
                time.sleep(0.008)
            except:
                pass
            
            # print('刷新成功')
    def check_buttom1(self):
        if(self.buttom1_state == 1):
            self.buttom1_state=0
            return 1
        
    def check_buttom2(self):
        if(self.buttom2_state == 1):
            self.buttom2_state=0
            return 1
    def wr_bt1(self):
        global buttom_pal
        self.buttom1_state = 1
        print(10)
        buttom_pal=1
    def wr_bt2(self):
        global buttom_stop
        self.buttom2_state = 1
        print(20)
        buttom_stop=1



# ui的线程 因为是class 开启新线程采用如下方法
class TT(threading.Thread):
    def run(self) -> None:
        mainui()
T2 = TT()
T2.start()

#物理按钮线程  用来在自由辩论切换计时用   如果没有物理按钮也可以自己再弄一个软件按钮
def physical_but():
    m=0
    global phy_but
    print("phy_but coming")
    while(True):
        while(GPIO.input(gpio_buttom) == 1):
            m+=1
            print('buttom pushed',m)
            phy_but = 1
            time.sleep(0.3)
t_but  = threading.Thread(target=physical_but)
t_but.start()




#beeper需要物理外设 暂不启用
"""
def beeper():
    global  warn0,warn1,warn2,warn3,gpio_beeper
    print("beeper starting")
    while(True):
        if(warn0==1):
            print('warn 0')
            warn0=0
            GPIO.output(gpio_beeper,1)
            time.sleep(0.3)
            GPIO.output(gpio_beeper,0)
            
        elif(warn1==1):
            warn1=0
            GPIO.output(gpio_beeper,1)
            time.sleep(2)
            GPIO.output(gpio_beeper,0)
        if(warn2==1):
            warn2=0
            GPIO.output(gpio_beeper,1)
            time.sleep(0.3)
            GPIO.output(gpio_beeper,0)
        elif(warn3==1):
            warn3=0
            GPIO.output(gpio_beeper,1)
            time.sleep(2)
            GPIO.output(gpio_beeper,0)
        time.sleep(0.1)
t_beep  = threading.Thread(target=beeper)
t_beep.start()
"""

# print('ok')

#读取excel表格存储的时间数据
local_path = os.getcwd()
wb = load_workbook(local_path+"/时间.xlsx")

#超声波传感器物理外设  暂不启用
"""
def sensor(team,option="defalt"):
    d1 = distance1()
    d2 = distance2()
    listre=[d1,d2]
    if(option == "defalt"):
        if(d1<d1_origin):
            listre[0] = 1
        else:
            listre[0]=0
        if(d2<d2_origin):
            listre[1] = 1
        else:
            listre[1]=0

    return listre[team-1]
"""
    

#程序以0.1s为一个tick  此函数为tick刷新函数
def ticktime(option):
    global ticks1,ticks2,time_start,ticks_backup
    if(option == "check"):
        pass
    elif(option == "delta"):
        delta =0.1- time.time()+time_start
        if(delta<0):
            #以防程序 掉tick 补足时间差
            while(delta<0):
                print("warning!!!cant keep up")
                delta+=0.1
                ticks_backup+=1
        # print('sleep')
        return delta
    elif(option == "freeze"):
        pass

#每一个阶段的退出检查  包含暂停，停止，跳过等功能
def exit_check():
    global state,ticks1,ticks2,maxtick,beepflag1,beepflag2
    if(buttom_stop == 1):
        
        return False
    if(state in [1,3,5,9]):
        if(ticks1>=maxtick-300):
            if(beepflag1==0):
                warn0=1
                beepflag1=1
            if(ticks1>=maxtick and beepflag1==1):
                warn1=1
                beepflag1=2
                # print("left_out",ticks1)
    elif(state in [2,4,6,8]):
        if(ticks2>=maxtick-300):
            if(beepflag2==0):
                warn2=1
                beepflag2=1
            if(ticks2>=maxtick and beepflag2==1):
                warn3=1
                beepflag2 = 2
            
    if(state!=7 and (ticks1>maxtick or ticks2>maxtick)):
        
        return False
    
    #自由辩论
    elif(state==7):
        #1
        if(ticks1>=maxtick-300):
            if(beepflag1==0):
                warn0=1
                beepflag1=1
            if(ticks1>=maxtick and beepflag1==1):
                warn1=1
                beepflag1=2
                
        #2
        if(ticks2>=maxtick-300):
            if(beepflag2==0):
                warn2 =1
                beepflag2=1
            if(ticks2>=maxtick and beepflag2==1):
                warn3=1
                beepflag2 = 2
        if(ticks1>=maxtick and ticks2>=maxtick):
        
            return False
        
    return True


#程序内部读取每一阶段时长 也就是tick值
def load_option(state):
    return ticks_list[state-1]

#这个函数主要用于处理当用户按下暂停或跳过（停止）按钮时的情况
def pal_or_break():
    global buttom_pal,buttom_stop,time_start,state_txt
    if(buttom_pal==1):
        buttom_pal=0
        # state_txt =state_list[state]+'(暂停状态)'
        while(buttom_pal==0 and buttom_stop==0):
            # check_buttom()
            time.sleep(0.2)
            # print('等待再次暂停，buttom_pal:',buttom_pal)
        buttom_pal=0
        # state_txt = state_list[state]
        time_start=time.time()
    if(buttom_stop==1):
        buttom_stop=0
        return 1
    
    return 0

#将tick格式转化为标准分秒格式
def time_formater(sec_in):
    if(sec_in>600):
        return str(sec_in//600)+':'+str((sec_in%600)//10)+'.'+str(sec_in%10)+'s'
    else:
        return '0 :'+str((sec_in%600)//10)+'.'+str(sec_in%10)+'s'


state = 1
state_list = ['','左方一辩','右方一辩','左方二辩','右方二辩','左方三辩','右方三辩','自由辩论','右方四辩','左方四辩']
ticks_list=[]
warn0 = 0
warn1 =   0
warn2= 0
warn3   = 0

warnlist=[0,0,0,0]

ticks1=0
ticks2=0

maxtick=0

beepflag1=0
beepflag2=0

teamstate=1

buttom_pal=0
buttom_stop=0

ticks_backup = 0


#将excel表格内数据转化到数组里面去
for ws in wb:
    for val in range(len(ws['B2':'B6'])):
        cell = ws['B2':'B6'][val]
        ticks_list.append(cell[0].value)
        if(val!=3):
            ticks_list.append(ticks_list[-1])
    print(ticks_list)

time.sleep(0.2)


#主程序
for i in range(len(ticks_list)):
    buttom_pal = 1   #一开始时默认进入暂停状态，再次按下暂停按钮开始计时
    buttom_stop = 0

    warn0 = 0
    warn1 = 0
    warn2= 0
    warn3= 0

    print('state',state)
    ticks1=0
    ticks2=0
    maxtick= load_option(state)  #读取这一阶段的时间（tick值）
    warnlist=[0,0,0,0]
    beepflag1=0
    beepflag2=0
    state_txt = state_list[state]
    while(exit_check()):  #每一次循环时就检查软件按钮状态
        time_start = time.time()  #开始计时
        
        if(pal_or_break()==1):
            break
        # t1 = sensor(1)
        # t2 = sensor(2)
        if(state in [1,3,5,9]):         #1号队伍

            if(ticks1>=maxtick-300):
                if(beepflag1==0):
                    warn0=1
                    beepflag1=1
                if(ticks1>=maxtick and beepflag1==1):
                    warn1=1
                    beepflag1=2

            time1 =time_formater(maxtick - ticks1)     #刷新ui显示的时间
            # print(time1)
            ticks1= ticks1+1+ticks_backup
            time.sleep(ticktime("delta"))   #所有步骤执行结束后，进行的等待 以程序循环开始时间记
        elif(state in [2,4,6,8]):    #2号队伍

            if(ticks2>=maxtick-300):
                if(beepflag2==0):
                    warn2=1
                    beepflag2=1
                if(ticks2>=maxtick and beepflag2==1):
                    warn3=1
                    beepflag2 = 2

            time2 =time_formater(maxtick - ticks2)
            ticks2=ticks2+1+ticks_backup
            time.sleep(ticktime("delta"))   
        elif(state ==7):       #自由辩论
            print(teamstate)       #计时队伍编号
            if(phy_but  == 1):             
                phy_but=0
                teamstate=-teamstate         #当物理按钮按下时  切换队伍  teamstate==1就是1队 -1就是2队
            if(ticks1<=maxtick and teamstate ==  1):
                time1 =time_formater(maxtick - ticks1)
                ticks1 = ticks1+1+ticks_backup
            elif(ticks2<=maxtick and teamstate   ==  -1):
                time2 =time_formater(maxtick - ticks2)
                ticks2=ticks2+1+ticks_backup
            if(ticks1  == maxtick):
                teamstate = -1
            if(ticks2 == maxtick):
                teamstate  = 1
            ticks_backup = 0
            time.sleep(ticktime("delta"))
            
    state+=1   #一个阶段结束进入下一个阶段

GPIO.cleanup()
#清空树莓派端口




